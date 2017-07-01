import os
import numpy as np
from openpyxl import load_workbook

# 文件夹切换
if 'Myxlsxdata' not in os.listdir():
    os.mkdir('Myxlsxdata')
os.chdir('Myxlsxdata')

# 读取当前工作目录下的xlsx文件
wb = load_workbook('pandas_moresheet.xlsx')

# 　查看所有表名
print("表名：", wb.get_sheet_names())

# 通过表名称来选择工作表
ws = wb['豆瓣图书']
print("行数：", len(list(ws.rows)))
print("列数：", len(list(ws.columns)))

# 获取一行的数据
# 这里，我们通过篇设置min_row和max_row相同来实现只取一行
# 一般第一行为列名，我们打印出来
row_data = []
for row in ws.iter_rows(min_row=1, max_row=1, max_col=5):
    for cell in row:
        row_data.append(cell.value)
print("第一行（列名）：", row_data)

# 获取某一列的数据,假设为第二列
row_data = []
for col in ws.iter_cols(min_col=2, max_col=2, max_row=41):
    for cell in col:
        row_data.append(cell.value)
print("第二列：", row_data)

# 获取某区块的数据
# 通过上面的程序也能看出，只要设置好row和col的阀值就行了
# 假设获取2-3列，1-5行的数据
print("区域数据(1-5，2-3):")
min_col = 2
max_col = 3
min_row = 1
max_row = 5

areadata = np.matrix(np.zeros((max_row - min_row + 1, max_col - min_col + 1)), dtype=str)
for col in ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row):
    for cell in col:
        col_index = cell.col_idx  # 获取所在列数
        row_index = cell.row  # 获取所在行数
        areadata[row_index - min_row, col_index - min_col] = cell.value

print(areadata)
